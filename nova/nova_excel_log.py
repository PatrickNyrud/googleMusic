import openpyxl
import nova_item

class excel_log:
    def __init__(self, items):
        self.dest = "logs//excel_logs//tot_salg.xlsx"
        self.work_book = openpyxl.load_workbook(self.dest)
        self.work_sheet = self.work_book.active

        self.obj = nova_item.items_frame()

        self.column_len = self.obj.pic_price_file("logs//", "lager.txt")

        for x in items: 
            self.write_workbook(x)
        
        self.save_excel()
    
    def get_columns(self, item):
        for x in range(1, len(self.column_len) + 1):
            #Gets the value of cell x
            self.column_value = self.work_sheet.cell(row = 1, column = x).value 
            #Checks name at cell x if its == to item
            if self.column_value == item:
                #Returns cell position
                return x
                #return self.work_sheet.cell(row = 1, column = x).column returns actuall name like AH
            else:
                pass

    def write_workbook(self, item):
        self.get_column = self.get_columns(item)
        print self.get_column
        self.column = self.work_sheet.cell(row = 2, column = self.get_column)
        self.column.value = self.column.value + 1

    #Save excel file
    def save_excel(self) :  
        self.work_book.save(self.dest)


