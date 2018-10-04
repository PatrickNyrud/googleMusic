import openpyxl
import svea_item
import time

class excel_log_pole:
    def __init__(self, items):
        self.dest = "logs//excel_logs//excel_pole.xlsx"
        self.work_book = openpyxl.load_workbook(self.dest)
        self.work_sheet = self.work_book.active

        self.obj = svea_item.items_frame()
        
        self.column_len = self.obj.pic_price_file("logs//", "lager.txt")
        
        for x in items: 
            self.write_workbook(x)
        
        self.save_excel()
    
    def get_columns(self, item):
        for x in range(1, len(self.column_len) + 1):
            #Gets the value of cell x
            self.column_value = self.work_sheet.cell(row = 1, column = x).value 
            #Checks name at cell x if its == to item
            if self.column_value == item:
                #Returns cell position
                return x
                #return self.work_sheet.cell(row = 1, column = x).column returns actuall name like AH
            else:
                pass

    def write_workbook(self, item):
        self.get_column = self.get_columns(item)
        self.column = self.work_sheet.cell(row = 2, column = self.get_column)
        self.column.value = self.column.value + 1

    #Save excel file
    def save_excel(self):  
        self.work_book.save(self.dest)


class excel_log_linear:
    def __init__(self, items):
        #18 rows
        #Sale at 1134, Super 10 x1
        self.dest = "logs//excel_logs//excel_linear.xlsx"
        self.work_book = openpyxl.load_workbook(self.dest)
        self.work_sheet = self.work_book.active

        self.obj = svea_item.items_frame()
        
        self.column_len = self.obj.pic_price_file("logs//", "lager.txt")

        for x in items:
            self.write_workbook(x)

        self.save_excel()

    def write_workbook(self, item):
        self.row = self.get_rows()
        self.column = self.get_columns(item)

        try:
            self.write = self.work_sheet.cell(row = self.row, column = self.column)
            self.write.value = self.write.value + 1
        except:
            pass

    def get_rows(self): 
        self.now = time.strftime("%H")
        for x in range(1, 18):
            self.row_value = self.work_sheet.cell(row = x, column = 1).value
            if self.now == self.row_value[0:2]:
                return x
            else:
                pass
                
    def get_columns(self, item):
        for x in range(1, len(self.column_len) + 1):
            #Gets the value of cell x
            self.column_value = self.work_sheet.cell(row = 1, column = x).value 
            #Checks name at cell x if its == to item
            if self.column_value == item:
                #Returns cell position
                return x
            else:
                pass
    
    def save_excel(self):  
        self.work_book.save(self.dest)


